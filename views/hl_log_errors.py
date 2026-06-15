import csv
import re
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import streamlit as st

_ERROR_RE = re.compile(
    r"ERROR:\s*Barcode\s+(\S+)\s+not found!\s+Quantity\s*=\s*([\d.]+)",
    re.IGNORECASE,
)

_DATA_DIR = Path(__file__).resolve().parent.parent / "data"

VENUES = {
    "RS — Red Sands": _DATA_DIR / "catalogue_RS.csv",
    "NH — Newman Hotel": _DATA_DIR / "catalogue_NH.csv",
}


def _parse_log(text: str) -> list[tuple[str, float]]:
    rows = []
    for line in text.splitlines():
        m = _ERROR_RE.search(line)
        if m:
            barcode = m.group(1).strip()
            qty = float(m.group(2))
            rows.append((barcode, qty))
    return rows


def _load_catalogue(path: Path) -> dict[str, str]:
    """Return {normalised_barcode: product_name}.
    Strips leading zeros so '029147100244' and '29147100244' resolve to the same key.
    """
    text = path.read_text(encoding="utf-8", errors="replace")
    reader = csv.DictReader(StringIO(text))
    headers = [h.strip().lower() for h in (reader.fieldnames or [])]

    code_col = next((h for h in headers if h in ("code", "barcode", "codigo", "sku")), None)
    name_col = next((h for h in headers if h in ("name", "nombre", "product", "description")), None)

    if not code_col or not name_col:
        return {}

    catalogue: dict[str, str] = {}
    for row in reader:
        code_val = ""
        name_val = ""
        for k, v in row.items():
            if k.strip().lower() == code_col:
                code_val = (v or "").strip()
            if k.strip().lower() == name_col:
                name_val = (v or "").strip()
        if code_val:
            normalised = code_val.lstrip("0") or "0"
            catalogue[normalised] = name_val
    return catalogue


def _lookup_name(barcode: str, catalogue: dict[str, str]) -> str:
    if not catalogue:
        return ""
    return catalogue.get(barcode.lstrip("0") or "0", "")


def _build_xlsx(data: dict[str, dict], catalogue: dict[str, str]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Not Found Barcodes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    center = Alignment(horizontal="center")

    has_catalogue = bool(catalogue)
    headers = ["Barcode", "Total Quantity"] + (["Suggested Name"] if has_catalogue else [])
    col_widths = [20, 16] + ([40] if has_catalogue else [])

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    for row_idx, (barcode, info) in enumerate(
        sorted(data.items(), key=lambda x: x[0]), start=2
    ):
        ws.cell(row=row_idx, column=1, value=barcode).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2, value=info["qty"]).alignment = center
        if has_catalogue:
            suggested = _lookup_name(barcode, catalogue)
            ws.cell(row=row_idx, column=3, value=suggested).alignment = Alignment(horizontal="left")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render():
    st.header("🔍 H&L Log Error Report")
    st.caption(
        "Upload one or more H&L import log files (.txt or .log) to extract barcodes "
        "that returned 'not found' errors, and download them as an Excel report."
    )

    venue = st.radio("Venue", list(VENUES.keys()), horizontal=True, key="hl_venue")
    catalogue = _load_catalogue(VENUES[venue])

    uploaded_files = st.file_uploader(
        "Upload log files (.txt or .log)",
        type=["txt", "log"],
        accept_multiple_files=True,
        key="hl_log_uploader",
    )

    if not uploaded_files:
        st.info("Upload one or more log files to continue.")
        return

    aggregated: dict[str, dict] = {}
    file_errors: list[str] = []
    total_raw = 0

    for f in uploaded_files:
        try:
            text = f.read().decode("utf-8", errors="replace")
            rows = _parse_log(text)
            total_raw += len(rows)
            for barcode, qty in rows:
                if barcode not in aggregated:
                    aggregated[barcode] = {"qty": 0.0}
                aggregated[barcode]["qty"] += qty
        except Exception as e:
            file_errors.append(f"**{f.name}**: {e}")

    for err in file_errors:
        st.error(err)

    if not aggregated:
        st.warning("No 'not found' errors detected in the uploaded files.")
        return

    matched = sum(1 for b in aggregated if _lookup_name(b, catalogue))
    st.success(
        f"Found **{total_raw}** error lines across {len(uploaded_files)} file(s) — "
        f"**{len(aggregated)}** unique barcodes. **{matched}** matched in {venue} catalogue."
    )

    xlsx_bytes = _build_xlsx(aggregated, catalogue)

    st.download_button(
        label="⬇️ Download report (.xlsx)",
        data=xlsx_bytes,
        file_name="hl_not_found_barcodes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
